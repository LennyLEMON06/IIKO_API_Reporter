import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from tkcalendar import Calendar
import requests
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib3
from urllib3.exceptions import InsecureRequestWarning
import os
import tempfile
import json
from collections import defaultdict

# Отключаем предупреждения о сертификатах
urllib3.disable_warnings(InsecureRequestWarning)

class IikoOlapReporterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("IIKO Reporter")
        self.root.geometry("1200x800")

        # Список доступных баз
        # Загрузка баз из внешнего файла
        self.available_bases = self.load_bases_config()

        self.selected_bases = []  
        self.report_data = {}   # Данные для отчета "Выручка динамика"
        self.revenue_data = {}  # Данные для отчета "Выручка динамика"
        self.create_widgets()

    def load_bases_config(self):
        """Загружает конфигурацию баз из JSON-файла"""
        config_path = os.path.join(os.path.dirname(__file__), "bases_config.json")
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл конфигурации не найден: {config_path}")
            return {}
        except json.JSONDecodeError as e:
            messagebox.showerror("Ошибка", f"Ошибка в формате JSON: {e}")
            return {}

    def create_widgets(self):
        # Main container frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel - bases selection
        left_panel = ttk.Frame(main_frame)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Right panel - period selection and action buttons
        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, padx=5, pady=5)
        
        # Frame для параметров подключения (с кнопкой авторизации)
        connection_frame = ttk.LabelFrame(left_panel, text="Параметры подключения", padding=10)
        connection_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(connection_frame, text="Логин:").grid(row=0, column=0, sticky=tk.W)
        self.login_entry = ttk.Entry(connection_frame)
        self.login_entry.grid(row=0, column=1, sticky=tk.EW, padx=5)
        self.login_entry.insert(0, "login")
        
        ttk.Label(connection_frame, text="Пароль:").grid(row=1, column=0, sticky=tk.W)
        self.password_entry = ttk.Entry(connection_frame, show="*")
        self.password_entry.grid(row=1, column=1, sticky=tk.EW, padx=5)
        self.password_entry.insert(0, "password")
        
        # Кнопка авторизации внутри блока параметров подключения
        self.auth_button = ttk.Button(connection_frame, text="Авторизация", command=self.auth)
        self.auth_button.grid(row=0, column=2, rowspan=2, padx=10, sticky=tk.NS)
        
        # Frame для выбора баз
        bases_frame = ttk.LabelFrame(left_panel, text="Выбор баз", padding=10)
        bases_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Кнопка "Выбрать все"
        select_all_frame = ttk.Frame(bases_frame)
        select_all_frame.pack(fill=tk.X, pady=5)
        
        self.select_all_var = tk.BooleanVar()
        self.select_all_btn = ttk.Checkbutton(
            select_all_frame, 
            text="Выбрать все", 
            variable=self.select_all_var,
            command=self.toggle_select_all
        )
        self.select_all_btn.pack(side=tk.LEFT, padx=5)
        
        # Список баз с чекбоксами
        self.base_vars = {}
        bases_list_frame = ttk.Frame(bases_frame)
        bases_list_frame.pack(fill=tk.BOTH, expand=True)
        
        # Создаем Canvas и Scrollbar для списка баз
        canvas = tk.Canvas(bases_list_frame)
        scrollbar = ttk.Scrollbar(bases_list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for base_name in self.available_bases:
            var = tk.BooleanVar()
            chk = ttk.Checkbutton(scrollable_frame, text=base_name, variable=var)
            chk.pack(anchor=tk.W, padx=5, pady=2)
            self.base_vars[base_name] = var
        
        # Frame для выбора периода (правая панель)
        period_frame = ttk.LabelFrame(right_panel, text="Выбор периода", padding=10)
        period_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Выпадающий список для быстрого выбора периода
        period_select_frame = ttk.Frame(period_frame)
        period_select_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(period_select_frame, text="Период:").pack(side=tk.LEFT, padx=5)
        
        self.period_var = tk.StringVar()
        self.period_combobox = ttk.Combobox(
            period_select_frame,
            textvariable=self.period_var,
            state="readonly"
        )
        self.period_combobox['values'] = [
            "Сегодня",
            "Вчера",
            "Текущая неделя",
            "Прошлая неделя",
            "Текущий месяц",
            "Прошлый месяц",
            "Текущий год",
            "Прошлый год",
            "Другой..."
        ]
        self.period_combobox.current(4)  # По умолчанию "Текущий месяц"
        self.period_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.period_combobox.bind("<<ComboboxSelected>>", self.update_period)
        
        # Календарь для выбора диапазона дат
        self.calendar_frame = ttk.Frame(period_frame)
        self.calendar_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        ttk.Label(self.calendar_frame, text="С:").grid(row=0, column=0, sticky=tk.W)
        self.cal_start = Calendar(
            self.calendar_frame, 
            selectmode='day',
            date_pattern='dd.MM.yyyy',
            locale='ru_RU'
        )
        self.cal_start.grid(row=1, column=0, padx=5, sticky=tk.NSEW)
        
        ttk.Label(self.calendar_frame, text="По:").grid(row=0, column=1, sticky=tk.W)
        self.cal_end = Calendar(
            self.calendar_frame, 
            selectmode='day',
            date_pattern='dd.MM.yyyy',
            locale='ru_RU'
        )
        self.cal_end.grid(row=1, column=1, padx=5, sticky=tk.NSEW)
        
        # Кнопки управления (под блоком выбора периода)
        button_frame = ttk.Frame(right_panel)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        # Кнопки управления (под блоком выбора периода)
        button_frame = ttk.Frame(right_panel)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        # Контейнер для всех кнопок в сетке
        grid_frame = ttk.Frame(button_frame)
        grid_frame.pack(fill=tk.X)

        # Настройка пропорций колонок
        grid_frame.columnconfigure(0, weight=1, uniform="buttons")
        grid_frame.columnconfigure(1, weight=1, uniform="buttons")

        # Первая строка: OLAP отчеты
        self.report_button = ttk.Button(
            grid_frame,
            text="Получить отчет OLAP-Планы", 
            command=self.get_report, 
            state=tk.DISABLED
        )
        self.report_button.grid(row=0, column=0, padx=5, pady=2, sticky="ew")

        self.export_button = ttk.Button(
            grid_frame,
            text="Экспорт Планы в Excel", 
            command=self.export_to_excel, 
            state=tk.DISABLED
        )
        self.export_button.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        # Вторая строка: Выручка
        self.revenue_button = ttk.Button(
            grid_frame,
            text="Получить отчет Выручка для динамики", 
            command=self.get_revenue_report, 
            state=tk.DISABLED
        )
        self.revenue_button.grid(row=1, column=0, padx=5, pady=2, sticky="ew")

        self.export_revenue_button = ttk.Button(
            grid_frame,
            text="Экспорт Выручка для динамики в Excel", 
            command=self.export_revenue_to_excel, 
            state=tk.DISABLED
        )
        self.export_revenue_button.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        # Третья строка: Акты списания (на всю ширину)
        self.writeoff_button = ttk.Button(
            grid_frame,
            text="Получить акты списания", 
            command=self.get_writeoff_report, 
            state=tk.DISABLED
        )
        self.writeoff_button.grid(row=2, column=0, columnspan=2, padx=5, pady=2, sticky="ew")
        
        # Лог событий (внизу окна)
        log_frame = ttk.LabelFrame(self.root, text="Лог", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = tk.Text(log_frame, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Инициализация дат
        self.update_period()

    def toggle_select_all(self):
        """Выбирает или снимает выбор со всех баз"""
        select_all = self.select_all_var.get()
        for var in self.base_vars.values():
            var.set(select_all)

    def get_selected_bases(self):
        """Возвращает список выбранных баз"""
        return [base_name for base_name, var in self.base_vars.items() if var.get()]

    def update_period(self, event=None):
        """Обновляет даты в календаре в зависимости от выбранного периода"""
        today = datetime.now()
        period = self.period_combobox.get()
        
        if period == "Сегодня":
            start_date = today
            end_date = today
        elif period == "Вчера":
            start_date = today - timedelta(days=1)
            end_date = today - timedelta(days=1)
        elif period == "Текущая неделя":
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == "Прошлая неделя":
            start_date = today - timedelta(days=today.weekday() + 7)
            end_date = start_date + timedelta(days=6)
        elif period == "Текущий месяц":
            start_date = today.replace(day=1)
            end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        elif period == "Прошлый месяц":
            start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
            end_date = today.replace(day=1) - timedelta(days=1)
        elif period == "Текущий год":
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        elif period == "Прошлый год":
            start_date = today.replace(year=today.year-1, month=1, day=1)
            end_date = today.replace(year=today.year-1, month=12, day=31)
        else:  # "Другой..."
            return
        
        self.cal_start.selection_set(start_date)
        self.cal_end.selection_set(end_date)

    def log_message(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def auth(self):
        self.log_message("Попытка авторизации...")
        self.report_button.config(state=tk.NORMAL)
        self.revenue_button.config(state=tk.NORMAL)
        self.writeoff_button.config(state=tk.NORMAL)
        self.log_message("✅ Авторизация успешна")

    def get_report(self):
        try:
            selected_bases = self.get_selected_bases()
            if not selected_bases:
                messagebox.showwarning("Ошибка", "Выберите хотя бы одну базу")
                return
            start_date, end_date = self.get_selected_dates()
            self.log_message(f"Загрузка отчета за период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            self.report_data = {}
            login = self.login_entry.get()
            password = self.password_entry.get()
            for base_name in selected_bases:
                self.log_message(f"Получение данных из базы: {base_name}...")
                base_info = self.available_bases[base_name]
                reporter = IikoOlapReporter(base_info["url"], login, password, base_info["preset_id"])
                if reporter.auth():
                    data = reporter.get_olap_report(start_date, end_date)
                    if data:
                        self.log_message(f"Получены данные из {base_name}")
                        # Добавляем лог структуры данных для отладки
                        self.log_message(f"Структура данных: {str(type(data))}, keys: {str(data.keys()) if isinstance(data, dict) else 'not dict'}")
                        self.log_message(f"Структура данных {base_name}: {str(data)[:500]}")
                        # Сохранить данные для дальнейшего анализа
                        with open(f"{base_name}_data.json", "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=4)
                        normalized_data = self.normalize_report_data(data)
                        if normalized_data:
                            self.report_data[base_name] = normalized_data
                            self.log_message(f"✅ Данные из {base_name} успешно обработаны (записей: {len(normalized_data)})")
                        else:
                            self.log_message(f"❌ Не удалось нормализовать данные из {base_name}")
                    else:
                        self.log_message(f"❌ Не удалось получить данные из {base_name}")
                else:
                    self.log_message(f"❌ Ошибка авторизации в базе {base_name}")
            if self.report_data:
                self.export_button.config(state=tk.NORMAL)
                self.log_message("✅ Все данные успешно загружены")
            else:
                self.log_message("❌ Не удалось получить данные ни из одной базы")
        except Exception as e:
            self.log_message(f"❌ Ошибка: {str(e)}")
            import traceback
            self.log_message(traceback.format_exc())

    def get_revenue_report(self):
        try:
            selected_bases = self.get_selected_bases()
            if not selected_bases:
                messagebox.showwarning("Ошибка", "Выберите хотя бы одну базу")
                return
            start_date, end_date = self.get_selected_dates()
            self.log_message(f"Загрузка отчета 'Выручка динамика' за период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            self.revenue_data = {}
            login = self.login_entry.get()
            password = self.password_entry.get()
            for base_name in selected_bases:
                self.log_message(f"Получение данных 'Выручка динамика' из базы: {base_name}...")
                base_info = self.available_bases[base_name]
                reporter = IikoRevenueReporter(base_info["url"], login, password, base_info["revenue_preset_id"])
                if reporter.auth():
                    data = reporter.get_olap_report(start_date, end_date)
                    if data:
                        self.log_message(f"Получены данные 'Выручка динамика' из {base_name}")
                        processed_data = reporter.process_report_data(data)
                        if processed_data and processed_data.get('has_data'):
                            self.revenue_data[base_name] = processed_data
                            self.log_message(f"✅ Данные 'Выручка динамика' из {base_name} успешно обработаны")
                        else:
                            self.log_message(f"❌ Не удалось обработать данные 'Выручка динамика' из {base_name}")
                    else:
                        self.log_message(f"❌ Не удалось получить данные 'Выручка динамика' из {base_name}")
                else:
                    self.log_message(f"❌ Ошибка авторизации в базе {base_name}")
            if self.revenue_data:
                self.export_revenue_button.config(state=tk.NORMAL)
                self.log_message("✅ Все данные 'Выручка динамика' успешно загружены")
            else:
                self.log_message("❌ Не удалось получить данные 'Выручка динамика' ни из одной базы")
        except Exception as e:
            self.log_message(f"❌ Ошибка: {str(e)}")
            import traceback
            self.log_message(traceback.format_exc())

    def get_selected_dates(self):
        """Возвращает выбранные даты в формате datetime"""
        start_date = datetime.strptime(self.cal_start.get_date(), '%d.%m.%Y')
        end_date = datetime.strptime(self.cal_end.get_date(), '%d.%m.%Y')
        return start_date, end_date

    def normalize_report_data(self, data):
        """Нормализует данные отчета для корректного создания DataFrame"""
        if isinstance(data, dict):
            # Проверяем наличие ключей с данными
            for key in ['records', 'data', 'report']:
                if key in data and isinstance(data[key], list):
                    return data[key]
            # Если не нашли массив, проверяем наличие других возможных ключей
            if 'result' in data and isinstance(data['result'], list):
                return data['result']
            # Если ничего не найдено, возвращаем весь словарь как список из одного элемента
            return [data]
        elif isinstance(data, list):
            return data
        else:
            return []

    def export_to_excel(self):
        try:
            if not self.report_data:
                messagebox.showwarning("Ошибка", "Нет данных для экспорта")
                return
            
            start_date, end_date = self.get_selected_dates()
            start_date_str = start_date.strftime('%d.%m.%Y')
            end_date_str = end_date.strftime('%d.%m.%Y')
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            # Пробуем несколько путей для сохранения
            save_paths = [
                os.path.join(os.path.expanduser("~"), "Documents"),
                os.path.join(os.path.expanduser("~"), "Desktop"),
                tempfile.gettempdir()
            ]
            for save_dir in save_paths:
                try:
                    filename = os.path.join(save_dir, f"OLAP-Планы {start_date_str}-{end_date_str} ({current_date}).xlsx")
                    self.log_message(f"Попытка сохранения в: {filename}")
                    # Проверяем доступность директории
                    if not os.path.exists(save_dir):
                        os.makedirs(save_dir)
                    wb = Workbook()
                    if 'Sheet' in wb.sheetnames:
                        del wb['Sheet']
                    start_date, end_date = self.get_selected_dates()
                    for base_name, data in self.report_data.items():
                        if not data:
                            self.log_message(f"Нет данных для {base_name}")
                            continue
                        try:
                            ws = wb.create_sheet(title=base_name[:31])
                            # Заголовки
                            ws['A1'] = "Планы"
                            ws['A1'].font = Font(bold=True, size=14)
                            ws.merge_cells('A1:G1')
                            ws['A2'] = f"Название ресторана: {base_name}"
                            ws['A3'] = f"Период: с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
                            # Заголовки столбцов
                            headers = ["Группа", "Неделя", "День недели", "Итого", 
                                    "Гости", "Блюда", "Чеки"]
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=5, column=col, value=header).font = Font(bold=True)
                                ws.column_dimensions[get_column_letter(col)].width = 15
                            # Сортируем данные сначала по группе, затем по неделе и дню
                            sorted_data = sorted(data, key=lambda x: (
                                x.get('RestorauntGroup', ''),
                                int(x.get('WeekInMonthOpen', 1) or 1),  # Безопасное преобразование
                                int(x.get('DayOfWeekOpen', '1. Понедельник').split('.')[0] or 1)  # Безопасное преобразование
                            ))
                            # Заполняем данные
                            row_num = 6
                            current_group = None
                            current_week = None
                            group_start_row = 6
                            week_start_row = 6
                            group_data = []  # Для хранения данных группы
                            
                            # Обработка каждой записи
                            for record in sorted_data:
                                # Обработка для базы "Казань 1, 2, 3 СХ Железногорск Брянск"
                                if base_name == "Казань 1, 2, 3 СХ Железногорск Брянск":
                                    # Для базы "Казань 1, 2, 3 СХ Железногорск Брянск" вычисляем недели и дни
                                    if 'WeekInMonthOpen' not in record or not record['WeekInMonthOpen']:
                                        record['WeekInMonthOpen'] = self._get_week_number(record['OpenDate.Typed'], start_date)
                                    if 'DayOfWeekOpen' not in record or not record['DayOfWeekOpen']:
                                        date = datetime.strptime(record['OpenDate.Typed'], '%Y-%m-%dT%H:%M:%S')
                                        day_num = date.weekday() + 1  # Нумерация дней недели с 1 (понедельник)
                                        day_name = self._get_day_name(day_num)
                                        record['DayOfWeekOpen'] = f"{day_num}. {day_name}"
                                
                                # Обработка для базы "СХ Орел" - проверяем наличие нужных полей
                                if base_name == "СХ Орел":
                                    # Для базы "СХ Орел" проверяем и корректируем поля
                                    if 'WeekInMonthOpen' not in record or not record['WeekInMonthOpen']:
                                        record['WeekInMonthOpen'] = '1'  # Установим значение по умолчанию
                                    if 'DayOfWeekOpen' not in record or not record['DayOfWeekOpen']:
                                        record['DayOfWeekOpen'] = '1. Понедельник'  # Установим значение по умолчанию
                                
                                group = record.get('RestorauntGroup', '')
                                week_num = int(record.get('WeekInMonthOpen', 1) or 1)
                                day_str = record.get('DayOfWeekOpen', '1. Понедельник')
                                
                                # Парсим день недели
                                day_parts = day_str.split('. ')
                                day_num = int(day_parts[0]) if day_parts and day_parts[0] else 1
                                day_name = day_parts[1] if len(day_parts) > 1 else "Понедельник"
                                
                                self.log_message(f"Обработка записи: Группа={group}, Неделя={week_num}, День={day_name}")
                                
                                # Если новая группа
                                if group != current_group:
                                    if current_group is not None:
                                        # Добавляем итог по последней неделе предыдущей группы
                                        self._add_week_total(ws, week_start_row, row_num-1, current_week)
                                        row_num += 1
                                        # Добавляем итог по всей группе
                                        self._add_group_total(ws, group_start_row, row_num-1, current_group)
                                        row_num += 1
                                    current_group = group
                                    current_week = week_num
                                    group_start_row = row_num
                                    week_start_row = row_num
                                    group_data = []
                                    # Записываем название группы
                                    ws.cell(row=row_num, column=1, value=group)
                                    row_num += 1
                                    week_start_row = row_num
                                # Если новая неделя в текущей группе
                                elif week_num != current_week:
                                    # Добавляем итог по предыдущей неделе
                                    self._add_week_total(ws, week_start_row, row_num-1, current_week)
                                    row_num += 1
                                    current_week = week_num
                                    week_start_row = row_num
                                # Записываем данные дня
                                ws.cell(row=row_num, column=1, value="")  # Пусто под названием группы
                                ws.cell(row=row_num, column=2, value=week_num)
                                ws.cell(row=row_num, column=3, value=f"{day_num}. {day_name}")
                                ws.cell(row=row_num, column=4, value=record.get('DishDiscountSumInt', 0))
                                ws.cell(row=row_num, column=5, value=record.get('GuestNum', 0))
                                ws.cell(row=row_num, column=6, value=record.get('DishAmountInt', 0))
                                ws.cell(row=row_num, column=7, value=record.get('UniqOrderId', 0))
                                row_num += 1
                                
                            # Добавляем итоги для последней недели и группы
                            if current_week is not None:
                                self._add_week_total(ws, week_start_row, row_num-1, current_week)
                                row_num += 1
                            if current_group is not None:
                                self._add_group_total(ws, group_start_row, row_num-1, current_group)
                        except Exception as e:
                            self.log_message(f"Ошибка в {base_name}: {str(e)}")
                            continue
                    # Сохраняем файл
                    wb.save(filename)
                    self.log_message(f"✅ Отчет успешно сохранен в: {filename}")
                    # Открываем файл
                    if os.name == 'nt':
                        os.startfile(filename)
                    return  # Успешно сохранили
                except PermissionError as e:
                    self.log_message(f"Ошибка доступа: {str(e)}")
                    continue
                except Exception as e:
                    self.log_message(f"Ошибка: {str(e)}")
                    continue
            # Если все попытки неудачны
            self.log_message("❌ Не удалось сохранить отчет ни в одну из папок")
            messagebox.showerror("Ошибка", "Не удалось сохранить отчет. Проверьте права доступа.")
        except Exception as e:
            self.log_message(f"Критическая ошибка: {str(e)}")
            messagebox.showerror("Ошибка", f"Ошибка: {str(e)}")

    def export_revenue_to_excel(self):
        try:
            if not self.revenue_data:
                messagebox.showwarning("Ошибка", "Нет данных для экспорта")
                return
            
            start_date, end_date = self.get_selected_dates()
            start_date_str = start_date.strftime('%d.%m.%Y')
            end_date_str = end_date.strftime('%d.%m.%Y')
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            
            # Определяем количество групп для каждой базы
            base_groups_config = {
                "Курск Ленина ММ": 4,
                "Анапа ММ": 1,
                "Курчатов CХ и Пекарни": 2,
                "Белгород СХ": 4,
                "Казань 1, 2, 3 СХ Железногорск Брянск": 5,
                "СХ Орел": 4,
                "Старый Оскол СХ": 2,
                "ИП Лозовская Пекарни": 4,
                "ИП Касаткин": 13
            }
            
            save_paths = [
                os.path.join(os.path.expanduser("~"), "Documents"),
                os.path.join(os.path.expanduser("~"), "Desktop"),
                tempfile.gettempdir()
            ]
            
            for save_dir in save_paths:
                try:
                    filename = os.path.join(save_dir, f"OLAP-Выручка для динамики {start_date_str}-{end_date_str} ({current_date}).xlsx")
                    self.log_message(f"Попытка сохранения в: {filename}")
                    
                    if not os.path.exists(save_dir):
                        os.makedirs(save_dir)
                    
                    wb = Workbook()
                    if 'Sheet' in wb.sheetnames:
                        del wb['Sheet']
                    
                    start_date, end_date = self.get_selected_dates()
                    
                    for base_name, data in self.revenue_data.items():
                        try:
                            # Получаем количество групп для текущей базы
                            groups_count = base_groups_config.get(base_name, 4)  # По умолчанию 4 группы
                            
                            # Создаем новый лист с именем базы (обрезаем до 31 символа)
                            sheet_name = base_name[:31]
                            ws = wb.create_sheet(title=sheet_name)

                            # Стили
                            bold_font = Font(bold=True)
                            center_alignment = Alignment(horizontal='center')
                            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                        top=Side(style='thin'), bottom=Side(style='thin'))

                            # Заголовок отчета
                            ws.merge_cells('A1:G1')
                            ws['A1'] = f"Выручка для динамики - {base_name}"
                            ws['A1'].font = Font(bold=True, size=14)
                            ws['A1'].alignment = center_alignment

                            # Информация о ресторане и периоде
                            ws['A2'] = f"Название ресторана: {base_name}"
                            ws['A3'] = f"Период: с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"

                            # Основная таблица
                            current_row = 5
                            
                            # Первая строка с группами
                            ws.merge_cells('A5:B5')
                            ws['A5'] = "Группа"
                            ws['A5'].font = bold_font
                            
                            # Получаем группы из данных (не больше указанного количества)
                            all_groups = data['groups']
                            groups = all_groups[:groups_count]
                            
                            # Определяем колонки для групп (A,B уже заняты)
                            columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P'][:len(groups)]
                            # Добавляем колонку "Итого"
                            total_col = chr(ord('C') + len(groups)) if len(groups) < 13 else 'P'
                            columns.append(total_col)
                            
                            # Записываем названия групп
                            for i, group in enumerate(groups):
                                ws[f'{columns[i]}{current_row}'] = group
                                ws[f'{columns[i]}{current_row}'].font = bold_font
                            
                            # Колонка "Итого"
                            ws[f'{columns[-1]}{current_row}'] = "Итого"
                            ws[f'{columns[-1]}{current_row}'].font = bold_font
                            
                            current_row += 1

                            # Строка с категорией блюда
                            ws.merge_cells(f'A{current_row}:B{current_row}')
                            ws[f'A{current_row}'] = "Категория блюда"
                            ws[f'A{current_row}'].font = bold_font
                            
                            # Месяц (используем текущий месяц)
                            current_month = f"{start_date.month:02d} ({self._get_month_name(start_date.month)})"
                            for col in columns:
                                ws[f'{col}{current_row}'] = current_month
                            
                            current_row += 1

                            # Пустая строка с нулями
                            ws.merge_cells(f'A{current_row}:B{current_row}')
                            for col in columns:
                                ws[f'{col}{current_row}'] = 0.00
                            current_row += 1

                            # Заполняем данные по категориям
                            for category in data['categories']:
                                ws.merge_cells(f'A{current_row}:B{current_row}')
                                ws[f'A{current_row}'] = category
                                
                                category_total = 0.0
                                for i, group in enumerate(groups):
                                    value = data['data'][group].get(category, 0.0)
                                    ws[f'{columns[i]}{current_row}'] = value
                                    category_total += value
                                
                                ws[f'{columns[-1]}{current_row}'] = category_total
                                current_row += 1

                            # Итоговая строка
                            ws.merge_cells(f'A{current_row}:B{current_row}')
                            ws[f'A{current_row}'] = "Итого"
                            ws[f'A{current_row}'].font = bold_font
                            
                            for i, group in enumerate(groups):
                                group_total = sum(data['data'][group].values())
                                ws[f'{columns[i]}{current_row}'] = group_total
                            
                            grand_total = sum(sum(group.values()) for group in data['data'].values())
                            ws[f'{columns[-1]}{current_row}'] = grand_total

                            # Форматирование
                            for row in ws.iter_rows(min_row=5, max_row=current_row, min_col=1, max_col=len(columns)+2):
                                for cell in row:
                                    cell.border = border
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = '#,##0.00'
                                    if cell.value in ["Итого", *groups, "Группа", "Категория блюда"]:
                                        cell.font = bold_font

                            # Настройка ширины столбцов
                            for col in ['A', 'B'] + columns:
                                ws.column_dimensions[col].width = 15

                            self.log_message(f"✅ Лист для базы {base_name} создан (групп: {len(groups)})")
                            
                        except Exception as e:
                            self.log_message(f"❌ Ошибка при создании листа для базы {base_name}: {str(e)}")
                            continue
                    
                    # Сохраняем файл
                    wb.save(filename)
                    self.log_message(f"✅ Отчет 'Выручка динамика' успешно сохранен в: {filename}")
                    # Открываем файл
                    if os.name == 'nt':
                        os.startfile(filename)
                    return  # Успешно сохранили
                    
                except PermissionError as e:
                    self.log_message(f"Ошибка доступа: {str(e)}")
                    continue
                except Exception as e:
                    self.log_message(f"Ошибка: {str(e)}")
                    continue
            
            # Если все попытки неудачны
            self.log_message("❌ Не удалось сохранить отчет ни в одну из папок")
            messagebox.showerror("Ошибка", "Не удалось сохранить отчет. Проверьте права доступа.")
            
        except Exception as e:
            self.log_message(f"Критическая ошибка: {str(e)}")
            messagebox.showerror("Ошибка", f"Ошибка: {str(e)}")
    

    def _get_day_name(self, day_num):
        """Возвращает название дня недели по номеру"""
        days = {
            1: "Понедельник",
            2: "Вторник",
            3: "Среда",
            4: "Четверг",
            5: "Пятница",
            6: "Суббота",
            7: "Воскресенье"
        }
        return days.get(day_num, "")

    def _get_month_name(self, month_num):
        """Возвращает название месяца по номеру"""
        months = {
            1: "Январь",
            2: "Февраль",
            3: "Март",
            4: "Апрель",
            5: "Май",
            6: "Июнь",
            7: "Июль",
            8: "Август",
            9: "Сентябрь",
            10: "Октябрь",
            11: "Ноябрь",
            12: "Декабрь"
        }
        return months.get(month_num, "")

    def _get_week_number(self, date_str, start_date):
        """Вычисляет номер недели в месяце"""
        date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S')
        # Номер недели в месяце
        week_in_month = (date.day - 1) // 7 + 1
        return week_in_month

    def _add_group_total(self, ws, start_row, end_row, group_name):
        """Добавляет строку с итогами по всей группе"""
        total_row = end_row + 1
        # Собираем все строки группы (исключая итоговые строки недель)
        group_rows = []
        for row in range(start_row, end_row + 1):
            # Проверяем, что это не итоговая строка недели
            if not ws.cell(row=row, column=2).value or "всего" not in str(ws.cell(row=row, column=2).value):
                group_rows.append(row)
        # Формулы для суммирования только по дням (без итогов недель)
        for col, letter in zip(range(4, 8), ['D', 'E', 'F', 'G']):
            # Собираем диапазон только для строк с данными дней
            range_str = ",".join([f"{letter}{row}" for row in group_rows])
            formula = f"=SUM({range_str})"
            ws.cell(row=total_row, column=col, value=formula).font = Font(bold=True)
        # Записываем итог по группе
        ws.cell(row=total_row, column=1, value=f"{group_name} всего").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="")
        ws.cell(row=total_row, column=3, value="")

    def _add_week_total(self, ws, start_row, end_row, week_num):
        """Добавляет строку с итогами по неделе (без дня недели)"""
        total_row = end_row + 1
        # Формулы для суммирования
        for col, letter in zip(range(4, 8), ['D', 'E', 'F', 'G']):
            formula = f"=SUM({letter}{start_row}:{letter}{end_row})"
            ws.cell(row=total_row, column=col, value=formula).font = Font(bold=True)
        # Записываем итог без дня недели
        ws.cell(row=total_row, column=1, value="")  # Пусто под названием группы
        ws.cell(row=total_row, column=2, value=f"{week_num} всего").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="")  # Пусто для дня недели

    def get_writeoff_report(self):
        try:
            selected_bases = self.get_selected_bases()
            if not selected_bases:
                messagebox.showwarning("Ошибка", "Выберите хотя бы одну базу")
                return
            
            start_date, end_date = self.get_selected_dates()
            start_date_str = start_date.strftime('%d.%m.%Y')
            end_date_str = end_date.strftime('%d.%m.%Y')
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
                
            start_date, end_date = self.get_selected_dates()
            self.log_message(f"Загрузка актов списания за период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            
            login = self.login_entry.get()
            password = self.password_entry.get()
            
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
                
            for base_name in selected_bases:
                base_info = self.available_bases[base_name]
                reporter = WriteoffReporter(base_info["url"], login, password)
                if reporter.auth():
                    # Загрузка справочников
                    reporter.load_stores_cache()
                    reporter.load_accounts_cache()
                    reporter.load_conceptions_cache()
                    reporter.load_products_cache()
                    
                    # Получение актов списания
                    docs = reporter.fetch_writeoff_docs(start_date, end_date)
                    if docs:
                        # Создание листа Excel
                        ws = wb.create_sheet(title=base_name[:31])
                        self.populate_writeoff_sheet(ws, docs, reporter, base_name)
                        self.log_message(f"✅ Данные из {base_name} успешно загружены")
                    else:
                        self.log_message(f"❌ Не удалось получить данные из {base_name}")
                else:
                    self.log_message(f"❌ Ошибка авторизации в базе {base_name}")
            
            # Сохранение файла
            filename = f"Акты списания {start_date_str}-{end_date_str} ({current_date}).xlsx"
            wb.save(filename)
            self.log_message(f"✅ Отчет по актам списания сохранен в: {filename}")
            if os.name == 'nt':
                os.startfile(filename)
                
        except Exception as e:
            self.log_message(f"❌ Ошибка: {str(e)}")
            import traceback
            self.log_message(traceback.format_exc())

    def populate_writeoff_sheet(self, ws, docs, reporter, base_name):
        """Заполняет лист Excel данными об актах списания"""
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Акты списания: {base_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment

        # Заголовки столбцов
        headers = [
            "Дата документа", "Тип", "№", "Товары", 
            "Сумма, р.", "Проведен", "Склад", 
            "Концепция", "Комментарий", "Счет списания"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Данные
        row_num = 4
        for doc in docs:
            if not isinstance(doc, dict):
                continue
                
            try:
                # Обработка даты
                date_str = datetime.strptime(
                    doc['dateIncoming'], 
                    '%Y-%m-%dT%H:%M:%S.%f'
                ).strftime('%d.%m.%Y %H:%M')
            except ValueError:
                try:
                    date_str = datetime.strptime(
                        doc['dateIncoming'], 
                        '%Y-%m-%dT%H:%M'
                    ).strftime('%d.%m.%Y %H:%M')
                except ValueError:
                    date_str = doc['dateIncoming']
            
            # Получение данных документа
            doc_num = doc.get('documentNumber', '')
            status = reporter.get_status_name(doc.get('status', ''))
            comment = doc.get('comment', '')
            
            # Товары и сумма
            items_text = []
            total_sum = 0
            for item in doc.get('items', []):
                product_name = reporter.get_product_name(item.get('productId', ''))
                amount = item.get('amount', 0)
                cost = item.get('cost', 0) or 0
                total_sum += cost
                items_text.append(f"{product_name} x{amount}")
            
            # Получение названий
            warehouse = reporter.get_store_name(doc.get('storeId'))
            writeoff_account = reporter.get_account_name(doc.get('accountId'))
            concept = reporter.get_conception_name(doc.get('conceptionId'))
            
            # Запись в Excel
            ws.cell(row=row_num, column=1, value=date_str).border = border
            ws.cell(row=row_num, column=2, value="P").border = border
            ws.cell(row=row_num, column=3, value=doc_num).border = border
            ws.cell(row=row_num, column=4, value="\n".join(items_text)).border = border
            ws.cell(row=row_num, column=5, value=round(total_sum, 2)).border = border
            ws.cell(row=row_num, column=6, value=status).border = border
            ws.cell(row=row_num, column=7, value=warehouse).border = border
            ws.cell(row=row_num, column=8, value=concept).border = border
            ws.cell(row=row_num, column=9, value=comment).border = border
            ws.cell(row=row_num, column=10, value=writeoff_account).border = border
            
            row_num += 1


class IikoOlapReporter:
    def __init__(self, base_url, login, password, preset_id):
        self.base_url = base_url
        self.login = login
        self.password = password
        self.preset_id = preset_id  # ID пресета отчета
        self.token = None
        self.session = requests.Session()
        self.session.verify = False

    def auth(self):
        """Авторизация в системе"""
        auth_url = f"{self.base_url}/auth"
        password_hash = hashlib.sha1(self.password.encode()).hexdigest()
        data = {
            'login': self.login,
            'pass': password_hash
        }
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = self.session.post(auth_url, data=data, headers=headers)
        if response.status_code == 200:
            self.token = response.text.strip()
            return True
        return False

    def get_olap_report(self, date_from, date_to):
        """Получение OLAP-отчета"""
        if not self.token and not self.auth():
            return None
        date_from_str = date_from.strftime('%Y-%m-%dT00:00:00')
        date_to_str = (date_to + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00')
        url = f"{self.base_url}/v2/reports/olap/byPresetId/{self.preset_id}"
        params = {
            'key': self.token,
            'dateFrom': date_from_str,
            'dateTo': date_to_str
        }
        response = self.session.get(url, params=params)
        if response.status_code == 200:
            return response.json()
        return None


class IikoRevenueReporter:
    def __init__(self, base_url, login, password, preset_id):
        self.base_url = base_url
        self.login = login
        self.password = password
        self.preset_id = preset_id
        self.token = None
        self.session = requests.Session()
        self.session.verify = False

    def auth(self):
        """Авторизация в системе"""
        auth_url = f"{self.base_url}/auth"
        password_hash = hashlib.sha1(self.password.encode()).hexdigest()
        data = {
            'login': self.login,
            'pass': password_hash
        }
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = self.session.post(auth_url, data=data, headers=headers)
        if response.status_code == 200:
            self.token = response.text.strip()
            return True
        return False

    def get_olap_report(self, date_from, date_to):
        """Получение OLAP-отчета"""
        if not self.token and not self.auth():
            return None
        date_from_str = date_from.strftime('%Y-%m-%dT00:00:00')
        date_to_str = (date_to + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00')
        url = f"{self.base_url}/v2/reports/olap/byPresetId/{self.preset_id}"
        params = {
            'key': self.token,
            'dateFrom': date_from_str,
            'dateTo': date_to_str
        }
        response = self.session.get(url, params=params)
        if response.status_code == 200:
            return response.json()
        return None

    def safe_get(self, dictionary, key, default=""):
        """Безопасное получение значения из словаря"""
        value = dictionary.get(key, default)
        return value.strip() if isinstance(value, str) else str(value)

    def process_report_data(self, json_data):
        """Обработка данных отчета с улучшенной обработкой ошибок"""
        if not json_data or not isinstance(json_data, dict) or 'data' not in json_data:
            return None

        try:
            groups = set()
            categories = set()
            data_by_group_category = defaultdict(lambda: defaultdict(float))

            for item in json_data['data']:
                if not isinstance(item, dict):
                    continue

                # Безопасное извлечение значений
                group = self.safe_get(item, 'RestorauntGroup')
                category = self.safe_get(item, 'DishCategory') or self.safe_get(item, 'DishGroup')
                month = self.safe_get(item, 'Mounth')

                # Пропускаем записи без категории или группы
                if not group or not category:
                    continue

                # Очистка и нормализация категорий
                category = category.strip()
                if category.startswith(" "):
                    category = category[1:]

                groups.add(group)
                categories.add(category)

                try:
                    amount = float(item.get('DishDiscountSumInt', 0))
                except (ValueError, TypeError):
                    amount = 0.0

                data_by_group_category[group][category] += amount

            if not groups or not categories:
                return None

            # Преобразуем в списки и сортируем
            groups = sorted(groups)
            categories = sorted(categories)

            return {
                'groups': groups,
                'categories': categories,
                'data': data_by_group_category,
                'has_data': len(groups) > 0 and len(categories) > 0
            }

        except Exception as e:
            return None
        
class WriteoffReporter:
    def __init__(self, base_url, login, password):
        self.base_url = base_url
        self.login = login
        self.password = password
        self.token = None
        self.session = requests.Session()
        self.session.verify = False
        self.stores_cache = None
        self.accounts_cache = None
        self.conceptions_cache = None
        self.products_cache = None

    def auth(self):
        """Авторизация в системе"""
        auth_url = f"{self.base_url}/auth"
        password_hash = hashlib.sha1(self.password.encode()).hexdigest()
        data = {
            'login': self.login,
            'pass': password_hash
        }
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = self.session.post(auth_url, data=data, headers=headers)
        if response.status_code == 200:
            self.token = response.text.strip()
            return True
        return False

    def load_stores_cache(self):
        """Загружает справочник складов"""
        api_url = f"{self.base_url}/v2/entities/list"
        params = {
            'key': self.token,
            'rootType': 'Account',
            'includeDeleted': 'false'
        }
        try:
            response = self.session.get(api_url, params=params)
            if response.status_code == 200:
                self.stores_cache = {
                    str(acc["id"]): acc["name"]
                    for acc in response.json()
                    if isinstance(acc, dict) and 
                    acc.get("type") == "INVENTORY_ASSETS" and 
                    "id" in acc and "name" in acc
                }
        except Exception as e:
            print(f"Error loading stores: {str(e)}")

    def load_accounts_cache(self):
        """Загружает справочник счетов"""
        api_url = f"{self.base_url}/v2/entities/accounts/list"
        params = {'key': self.token, 'includeDeleted': 'false'}
        try:
            response = self.session.get(api_url, params=params)
            if response.status_code == 200:
                self.accounts_cache = {
                    str(acc["id"]): acc["name"]
                    for acc in response.json()
                    if isinstance(acc, dict) and "id" in acc and "name" in acc
                }
        except Exception as e:
            print(f"Error loading accounts: {str(e)}")

    def load_conceptions_cache(self):
        """Загружает справочник концепций"""
        api_url = f"{self.base_url}/v2/entities/list"
        params = {
            'key': self.token,
            'rootType': 'Conception',
            'includeDeleted': 'false'
        }
        try:
            response = self.session.get(api_url, params=params)
            if response.status_code == 200:
                self.conceptions_cache = {
                    str(item["id"]): item["name"]
                    for item in response.json()
                    if isinstance(item, dict) and 
                    item.get("rootType") == "Conception" and 
                    "id" in item and "name" in item
                }
        except Exception as e:
            print(f"Error loading conceptions: {str(e)}")

    def load_products_cache(self):
        """Загружает справочник товаров"""
        api_url = f"{self.base_url}/v2/entities/products/list"
        params = {'key': self.token}
        try:
            response = self.session.get(api_url, params=params)
            if response.status_code == 200:
                self.products_cache = {
                    product["id"]: product["name"]
                    for product in response.json()
                    if isinstance(product, dict) and 
                    "id" in product and "name" in product
                }
        except Exception as e:
            print(f"Error loading products: {str(e)}")

    def fetch_writeoff_docs(self, date_from, date_to):
        """Получает акты списания за период"""
        api_url = f"{self.base_url}/v2/documents/writeoff"
        params = {
            'key': self.token,
            'dateFrom': date_from.strftime('%Y-%m-%d'),
            'dateTo': date_to.strftime('%Y-%m-%d')
        }
        try:
            response = self.session.get(api_url, params=params)
            if response.status_code == 200:
                return response.json().get("response", [])
            return None
        except Exception as e:
            print(f"Error fetching writeoff docs: {str(e)}")
            return None

    def get_store_name(self, store_id):
        return self.stores_cache.get(store_id, 'Неизвестно') if self.stores_cache else 'Неизвестно'

    def get_account_name(self, account_id):
        return self.accounts_cache.get(account_id, 'Неизвестно') if self.accounts_cache else 'Неизвестно'

    def get_conception_name(self, conception_id):
        return self.conceptions_cache.get(conception_id, 'Без концепции') if self.conceptions_cache else 'Без концепции'

    def get_product_name(self, product_id):
        return self.products_cache.get(product_id, 'Неизвестно') if self.products_cache else 'Неизвестно'

    @staticmethod
    def get_status_name(status):
        return {'NEW': 'Новый', 'PROCESSED': 'Да', 'DELETED': 'Удалённый'}.get(status, status)


if __name__ == "__main__":
    root = tk.Tk()
    app = IikoOlapReporterGUI(root)
    root.mainloop()